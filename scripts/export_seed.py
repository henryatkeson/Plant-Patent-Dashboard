from __future__ import annotations

import json
import re
import shutil
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT.parent / "North American fruit and nut patents PBR and HortScience 2016-2026.xlsx"
OUT = ROOT / "data" / "plant_patents.json"


def cell_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return ""


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def source_kind(primary_source: str) -> str:
    source = primary_source.upper().strip()
    if source.startswith("USPPA"):
        return "Published plant application"
    if source.startswith("USPP"):
        return "Issued plant patent"
    if source.startswith("USPVPA") or source.startswith("USPVP"):
        return "USDA PVPA/PVP"
    if source.startswith("HORTSCIENCE") or source.startswith("ACTA"):
        return "Horticultural publication"
    if source.startswith("CA"):
        return "Canadian PBR"
    if source.startswith("MX"):
        return "Mexican PBR"
    return "Other"


def normalize_patent_number(primary_source: str) -> str:
    match = re.search(r"\bUSPP\s*([0-9,]+)\b", primary_source, re.I)
    if not match:
        return ""
    return "USPP" + match.group(1).replace(",", "")


def main() -> None:
    with tempfile.TemporaryDirectory() as temp_dir:
        workbook_copy = Path(temp_dir) / WORKBOOK.name
        shutil.copy2(WORKBOOK, workbook_copy)
        wb = load_workbook(workbook_copy, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    headers = [text(value) or f"Column {idx + 1}" for idx, value in enumerate(rows[0])]
    records: list[dict[str, Any]] = []

    for row_number, row in enumerate(rows[1:], start=2):
        item = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
        primary_source = text(item.get("Primary Source"))
        crop = text(item.get("Crop"))
        if not primary_source and not crop:
            continue

        patent_number = normalize_patent_number(primary_source)
        date = cell_date(item.get("Date"))
        record_id = patent_number or f"seed-row-{row_number}"
        title_bits = [text(item.get("Cultivar denomination")), crop]
        title = " - ".join(bit for bit in title_bits if bit)

        records.append(
            {
                "id": record_id,
                "source": "Workbook seed",
                "sourceKind": source_kind(primary_source),
                "primarySource": primary_source,
                "patentNumber": patent_number,
                "publicationNumber": primary_source if primary_source.upper().startswith("USPPA") else "",
                "date": date,
                "issueDate": date if patent_number else "",
                "title": title,
                "crop": crop,
                "cultivar": text(item.get("Cultivar denomination")),
                "tradeName": text(item.get("Trade name")),
                "status": text(item.get("IP Status")),
                "breeders": text(item.get("Breeder(s)")),
                "assignee": "",
                "inventors": text(item.get("Breeder(s)")),
                "list": text(item.get("List")),
                "notes": " | ".join(
                    part
                    for part in [
                        text(item.get("Other sources / Notes")),
                        text(item.get("Other sources / Notes 2")),
                    ]
                    if part
                ),
                "sourceUrl": "",
                "detailText": "",
                "updatedAt": datetime.now(timezone.utc).isoformat(),
            }
        )

    records.sort(key=lambda row: row.get("date") or "", reverse=True)
    payload = {
        "metadata": {
            "title": "Fruit, Tree Nut, and Vegetable Plant Patent Dashboard",
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "recordCount": len(records),
            "sources": [
                "Seeded from the local Excel workbook",
                "Daily updater checks USPTO Official Gazette public plant-patent pages",
            ],
            "publicDataNote": (
                "The dashboard tracks public issued grants and public/published records. "
                "Unpublished filings inside Patent Center are not public dashboard data."
            ),
        },
        "records": records,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {len(records)} records to {OUT}")


if __name__ == "__main__":
    main()
