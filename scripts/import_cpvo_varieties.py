#!/usr/bin/env python3
"""Normalize CPVO Variety Finder workbook exports for the dashboard."""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK_GLOB = "*CPOV Varieties.xlsx"
OUTPUT_PATH = ROOT / "data" / "cpvo_varieties.json"

REGISTER_LABELS = {
    "PBR": "Plant breeders' rights",
    "PLP": "Plant patent",
    "NLI": "National list",
    "FRU": "Fruit register",
    "COM": "Commercial register",
    "ZZZ": "Other register",
}

REGISTER_GROUPS = {
    "PBR": "Protected IP",
    "PLP": "Protected IP",
    "NLI": "Official listing",
    "FRU": "Official listing",
    "COM": "Commercial listing",
    "ZZZ": "Other / unclear",
}


def clean_text(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text).strip()


def parse_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()

    text = clean_text(value)
    if not text:
        return ""

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            parsed = dt.datetime.strptime(text, fmt).date()
            if parsed > dt.date.today():
                return ""
            return parsed.isoformat()
        except ValueError:
            pass
    return text


def infer_crop(species_class: str, species_latin_name: str) -> str:
    species_class = species_class.upper()
    latin = species_latin_name.lower()

    if "fragaria" in latin or species_class in {"FRAGA", "FRPOT"}:
        return "Strawberry"
    if "vitis" in latin or species_class == "VITIS":
        return "Grape"
    if "vaccinium macrocarpon" in latin or "vaccinium oxycoccos" in latin:
        return "Cranberry"
    if "vaccinium vitis-idaea" in latin:
        return "Lingonberry"
    if "vaccinium" in latin or species_class == "VACCI":
        return "Blueberry"
    if "malus" in latin or species_class == "MALUS":
        return "Apple"
    if "cydonia" in latin or species_class == "CYDON":
        return "Quince"
    if "cydolus" in latin or species_class == "CYDOL":
        return "Apple/Quince Hybrid"
    if "rubus" in latin or species_class == "RUBUS":
        if any(token in latin for token in ["subg. rubus", "eubatus", "fruticosus", "ursinus"]):
            return "Blackberry"
        if any(token in latin for token in ["idaeus", "occidentalis", "arcticus", "chamaemorus", "coreanus", "niveus"]):
            return "Raspberry"
        return "Rubus"
    if "prunus" in latin or species_class.startswith("CL6"):
        if "nucipersica" in latin:
            return "Nectarine"
        if "persica" in latin:
            return "Peach"
        if "dulcis" in latin or "amygdalus" in latin:
            return "Almond"
        if "armeniaca" in latin:
            return "Apricot"
        if "avium" in latin:
            return "Cherry-Sweet"
        if "cerasus" in latin or "fruticosa" in latin or "gondouinii" in latin:
            return "Cherry-Tart"
        if any(token in latin for token in ["domestica", "salicina", "cerasifera", "mume", "limeixing"]):
            return "Plum"
        return "Prunus"
    if "citrus" in latin or "eremocitrus" in latin or species_class in {"CITRU", "EREMC"}:
        return "Citrus"
    if "pyrus" in latin or species_class == "PYRUS":
        return "Pear"
    if "olea" in latin or species_class == "OLEAA":
        return "Olive"
    if "juglans" in latin or species_class == "JUGLA":
        return "Walnut"
    if "actinidia" in latin or species_class == "ACTIN":
        return "Kiwifruit"
    if "musa" in latin or "ensete" in latin or species_class in {"MUSAA", "ENSET"}:
        return "Banana"
    if "theobroma" in latin or species_class == "THEOB":
        return "Cacao"
    if "corylus" in latin or species_class == "CRYLS":
        return "Hazelnut"
    if "persea" in latin or species_class == "PERSE":
        return "Avocado"
    if "mangifera" in latin or species_class == "MANGI":
        return "Mango"
    if "pistacia" in latin or species_class == "PISTA":
        return "Pistachio"
    if "ananas" in latin or species_class == "ANANA":
        return "Pineapple"
    if "carica" in latin or species_class == "CARIC":
        return "Papaya"
    if "carya" in latin or species_class == "CARYA":
        return "Pecan"
    if "ribes nigrum" in latin:
        return "Currant-Black"
    if "ribes uva-crispa" in latin:
        return "Gooseberry"
    if "ribes" in latin or species_class == "RIBES":
        return "Currant"
    if "ficus" in latin or species_class == "FICUS":
        return "Fig"
    if "punica" in latin or species_class == "PUNIC":
        return "Pomegranate"
    if "passiflora" in latin or species_class == "PASSI":
        return "Passion Fruit"
    if "annona cherimola" in latin:
        return "Cherimoya"
    if "annona muricata" in latin:
        return "Soursop"
    if "annona" in latin or species_class == "ANNON":
        return "Annona"
    if "psidium" in latin or species_class == "PSIDI":
        return "Guava"
    if "cyperus" in latin or species_class == "CYPER":
        return "Other CPVO"
    return species_class or "Unclassified"


def stable_id(row: dict[str, str]) -> str:
    key = "|".join(
        [
            row.get("Denomination", ""),
            row.get("Species latin name", ""),
            row.get("Species class", ""),
            row.get("Country", ""),
            row.get("Register type", ""),
            row.get("Application number", ""),
            row.get("Application date", ""),
            row.get("Grant/registration date", ""),
        ]
    )
    return "CPVO-" + hashlib.sha1(key.encode("utf-8")).hexdigest()[:12].upper()


def normalize_row(row: dict[str, str], workbook_name: str) -> dict[str, Any]:
    register_type = row.get("Register type", "").upper()
    species_class = row.get("Species class", "").upper()
    application_date = parse_date(row.get("Application date"))
    grant_date = parse_date(row.get("Grant/registration date"))
    date = grant_date or application_date
    register_label = REGISTER_LABELS.get(register_type, register_type or "CPVO register")
    register_group = REGISTER_GROUPS.get(register_type, "Other / unclear")
    denomination = row.get("Denomination", "") or "Denomination not available"
    species_latin_name = row.get("Species latin name", "")
    country = row.get("Country", "")
    application_number = row.get("Application number", "")

    record = {
        "id": stable_id(row),
        "sourceKind": f"CPVO {register_label}",
        "primarySource": application_number or f"{register_type} {country}".strip(),
        "date": date,
        "applicationDate": application_date,
        "grantDate": grant_date,
        "crop": infer_crop(species_class, species_latin_name),
        "cultivar": denomination,
        "status": row.get("Variety status", "").title(),
        "denominationStatus": row.get("Denomination status", "").title(),
        "applicationNumber": application_number,
        "breeders": row.get("Breeder", ""),
        "breederReference": row.get("Breeder\u2018s reference", ""),
        "registerType": register_type,
        "registerLabel": register_label,
        "registerGroup": register_group,
        "country": country,
        "speciesClass": species_class,
        "speciesLatinName": species_latin_name,
        "denominationNature": row.get("Denomination nature", ""),
    }
    return {key: value for key, value in record.items() if value not in ("", None)}


def load_records(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [clean_text(cell) for cell in next(rows)]
    records: list[dict[str, Any]] = []
    for row in rows:
        raw = {headers[index]: clean_text(value) for index, value in enumerate(row) if index < len(headers)}
        if not any(raw.values()):
            continue
        records.append(normalize_row(raw, workbook_path.name))
    return records


def discover_workbooks() -> list[Path]:
    return sorted(ROOT.parent.glob(DEFAULT_WORKBOOK_GLOB), key=lambda path: path.name.lower())


def load_all_records(workbook_paths: list[Path], sheet_name: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int]:
    records_by_id: dict[str, dict[str, Any]] = {}
    source_workbooks: list[dict[str, Any]] = []
    raw_count = 0

    for workbook_path in workbook_paths:
        records = load_records(workbook_path, sheet_name)
        raw_count += len(records)
        source_workbooks.append({"name": workbook_path.name, "records": len(records)})
        for record in records:
            existing = records_by_id.get(record["id"])
            if not existing:
                records_by_id[record["id"]] = record
                continue
            sources = set(existing.get("duplicateSourceWorkbooks", []))
            sources.add(existing.get("sourceWorkbook", ""))
            sources.add(record.get("sourceWorkbook", ""))
            existing["duplicateSourceWorkbooks"] = sorted(source for source in sources if source)

    return list(records_by_id.values()), source_workbooks, raw_count


def write_output(records: list[dict[str, Any]], source_workbooks: list[dict[str, Any]], raw_count: int, sheet_name: str) -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    register_counts: dict[str, int] = {}
    crop_counts: dict[str, int] = {}
    species_class_counts: dict[str, int] = {}
    for row in records:
        register_counts[row["registerType"]] = register_counts.get(row["registerType"], 0) + 1
        crop_counts[row["crop"]] = crop_counts.get(row["crop"], 0) + 1
        species_class_counts[row["speciesClass"]] = species_class_counts.get(row["speciesClass"], 0) + 1

    payload = {
        "metadata": {
            "title": "CPVO Variety Finder Records",
            "sourceWorkbooks": source_workbooks,
            "sourceSheet": sheet_name,
            "sourceUrl": "https://online.plantvarieties.eu/",
            "generatedAt": dt.datetime.now(dt.timezone.utc).isoformat(),
            "rawRecordCount": raw_count,
            "recordCount": len(records),
            "duplicateCount": raw_count - len(records),
            "registerCounts": dict(sorted(register_counts.items())),
            "cropCounts": dict(sorted(crop_counts.items())),
            "speciesClassCounts": dict(sorted(species_class_counts.items())),
        },
        "records": records,
    }
    OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, action="append", default=[])
    parser.add_argument("--sheet", default="Varieties")
    args = parser.parse_args()

    workbook_paths = [path.expanduser().resolve() for path in args.workbook] if args.workbook else discover_workbooks()
    if not workbook_paths:
        raise FileNotFoundError(f"Could not find any workbooks matching {DEFAULT_WORKBOOK_GLOB} in {ROOT.parent}")
    missing = [path for path in workbook_paths if not path.exists()]
    if missing:
        raise FileNotFoundError("Could not find workbook(s): " + ", ".join(str(path) for path in missing))

    records, source_workbooks, raw_count = load_all_records(workbook_paths, args.sheet)
    records = sorted(records, key=lambda row: (row.get("date", ""), row.get("cultivar", "")), reverse=True)
    write_output(records, source_workbooks, raw_count, args.sheet)
    print(f"Read {raw_count:,} CPVO rows from {len(workbook_paths)} workbook(s).")
    print(f"Wrote {OUTPUT_PATH} with {len(records):,} unique CPVO records.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
